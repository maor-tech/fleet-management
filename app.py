"""
Fleet Management System — Noamer
מערכת ניהול צי רכבים
"""

import streamlit as st
import pandas as pd
import json
import os
import hashlib
import base64
from datetime import datetime, date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Fleet Management — Noamer",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Constants ──────────────────────────────────────────────────────────────
COMPANIES  = ["Shlomo", "Avis", "Blue Sky", "Albar"]
CATEGORIES = ["Pickup Truck 4x4","Mid-Size SUV","Full-Size SUV",
               "Full-Size SUV Plus","Cargo/Van","1/2 Ton Truck","Sedan","Other"]
PRICE_USD  = {"Pickup Truck 4x4":4139,"Mid-Size SUV":2475,"Full-Size SUV":3899,
               "Full-Size SUV Plus":5424,"Cargo/Van":3729,"1/2 Ton Truck":5729,"Sedan":2034,"Other":2475}
PRICE_ILS  = {"Pickup Truck 4x4":12210,"Mid-Size SUV":7300,"Full-Size SUV":11500,
               "Full-Size SUV Plus":16000,"Cargo/Van":11000,"1/2 Ton Truck":16900,"Sedan":6000,"Other":8000}
COSTS_ILS  = {
    "Pickup Truck 4x4":{"Avis":9435,"Albar":8360,"Shlomo":9900,"Blue Sky":8500},
    "Mid-Size SUV":    {"Avis":5000,"Albar":5032,"Shlomo":5250,"Blue Sky":None},
    "Full-Size SUV":   {"Avis":9750,"Albar":8810,"Shlomo":13200,"Blue Sky":8000},
    "1/2 Ton Truck":   {"Avis":None,"Albar":None,"Shlomo":None,"Blue Sky":13500},
    "Cargo/Van":       {"Avis":10440,"Albar":8760,"Shlomo":8350,"Blue Sky":9500},
}
CO_BG = {"Shlomo":"#E8F5E9","Avis":"#FCE4EC","Blue Sky":"#E3F2FD","Albar":"#FFF9E6"}
DB_FILE   = "vehicles_db.json"
SEED_FILE = "vehicles_clean.json"

# ── Auth ───────────────────────────────────────────────────────────────────
def verify_password(stored_hash: str, password: str) -> bool:
    try:
        decoded = base64.b64decode(stored_hash.encode())
        salt, key = decoded[:16], decoded[16:]
        new_key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 100000)
        return new_key == key
    except:
        return False

def get_users():
    return st.secrets.get("users", {})

def login_page():
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"]{background:#f0f4f8}
    .login-card{background:white;border-radius:16px;padding:2.5rem;
                max-width:420px;margin:6rem auto 0;
                box-shadow:0 4px 24px rgba(0,0,0,0.10)}
    .login-logo{text-align:center;margin-bottom:1.5rem}
    .login-logo h1{font-size:1.5rem;color:#1B2A4A;margin:0.5rem 0 0.2rem}
    .login-logo p{color:#6c757d;font-size:0.85rem;margin:0}
    </style>
    <div class="login-card">
      <div class="login-logo">
        <div style="font-size:2.5rem">🚗</div>
        <h1>Fleet Management</h1>
        <p>Noamer יעוץ והשקעות</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        st.markdown("<div style='max-width:420px;margin:0 auto'>", unsafe_allow_html=True)
        email    = st.text_input("אימייל · Email", placeholder="you@noamer.co")
        password = st.text_input("סיסמה · Password", type="password")
        submitted = st.form_submit_button("כניסה למערכת →", use_container_width=True, type="primary")
        st.markdown("</div>", unsafe_allow_html=True)

    if submitted:
        users = get_users()
        email_lower = email.strip().lower()
        matched = next((u for k, u in users.items() if u["email"].lower() == email_lower), None)
        if matched and verify_password(matched["password_hash"], password):
            st.session_state["user"] = {
                "email": matched["email"],
                "name":  matched["name"],
                "role":  matched["role"],
            }
            st.rerun()
        else:
            st.error("אימייל או סיסמה שגויים")

def logout():
    st.session_state.pop("user", None)
    st.rerun()

# ── Database ───────────────────────────────────────────────────────────────
def load_db():
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    if os.path.exists(SEED_FILE):
        with open(SEED_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        save_db(data)
        return data
    return []

def save_db(data):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def next_id(data):
    return max((v["id"] for v in data), default=0) + 1

# ── Category normalizer ────────────────────────────────────────────────────
def normalize_cat(raw):
    c = str(raw).upper().replace("\xa0"," ").strip()
    if any(x in c for x in ["PICK UP","HILUX","D-MAX","MUSSO","AMAROK","RANGER","SSANGYONG","ISUZU","ADVENTURE"]):
        return "Pickup Truck 4x4"
    if any(x in c for x in ["MINIVAN","VAN 9","VAN L2H","PANEL VAN","TRAFIC","STARIA","SCUDO","BERLINGO","DUBLO","COMBO","CARGO","DUCATO","BOXER"]):
        return "Cargo/Van"
    if any(x in c for x in ["SMALL SUV","ECLIPSE","KOLEOS","TUCSON","SORENTO","CROSSTREK","FORESTER","SPORTAGE","OUTLANDER","CX5","CX-5","TRAILBLAZER","KODIAQ","BAYON","ARONA","3008","KAMIQ"]):
        return "Mid-Size SUV"
    if any(x in c for x in ["LANDCRUISER","LAND CRUISER","FULL SIZE SUV PLUS","FULL-SIZE SUV PLUS"]):
        return "Full-Size SUV Plus"
    if any(x in c for x in ["TRAVERSE","FULL-SIZE SUV","FULL SIZE SUV","SUV AUTOMAT"]):
        return "Full-Size SUV"
    if any(x in c for x in ["RAM","1/2 TON","DODGE RAM"]):
        return "1/2 Ton Truck"
    if any(x in c for x in ["SEDAN","COROLLA","CAMRY"]):
        return "Sedan"
    return "Other"

# ── Supplier parser ────────────────────────────────────────────────────────
ALIASES = {
    "vehicle_num":["מס' רכב","מס רכב","vehicle #","vehicle_num","'מס רכב","'מס.1","רכב"],
    "start":      ["תאריך השכרה","תאריך תחילת חוזה","start","rental start","תאריך התחלה"],
    "contract":   ["מס' חוזה","מס חוזה","contract","'מס חוזה"],
    "model":      ["סוג רכב","דגם","model","קבוצת הרכב שם בחוזה","שם דגם","רכב.1"],
    "category":   ["קטגוריה","category","vehicle category","קבוצת הרכב"],
    "rate_ils":   ["מחיר","תעריף","rate","עלות","monthly rate"],
    "driver":     ["שם פרטי","נהג","driver","שם לקוח"],
    "branch":     ["סניף","branch","שם סניף"],
    "days":       ["ימים","days","ימים לחוזה"],
    "km":         ['ק"מ יציאה','km','ק"מ'],
}

def detect_col(cols, field):
    for alias in ALIASES.get(field, []):
        for col in cols:
            if alias.lower().replace("'","").replace(" ","") in str(col).lower().replace("'","").replace(" ","").replace("\xa0",""):
                return col
    return None

def detect_company(filename, sheet=""):
    t = (filename + " " + sheet).lower()
    if "שלמה" in t or "shlomo" in t: return "Shlomo"
    if "אלבר" in t or "albar" in t or "alber" in t: return "Albar"
    if "אוויס" in t or "avis" in t or "auto" in t: return "Avis"
    if "בלו סקיי" in t or "blue sky" in t: return "Blue Sky"
    return None

def parse_file(uploaded, company_override=None):
    results, errors = [], []
    try:
        sheets = pd.read_excel(uploaded, sheet_name=None, dtype=str)
    except:
        try:
            sheets = pd.read_excel(uploaded, sheet_name=None, dtype=str, engine="xlrd")
        except Exception as e:
            return [], [f"שגיאה בקריאת הקובץ: {e}"]

    for sname, df in sheets.items():
        if df.empty or len(df) < 2: continue
        df.columns = [str(c).replace("\xa0"," ").strip() for c in df.columns]
        company = company_override or detect_company(uploaded.name, sname) or sname
        cm = {f: detect_col(df.columns, f) for f in ALIASES}
        if not cm["vehicle_num"]:
            errors.append(f"גיליון '{sname}': לא נמצאה עמודת מס׳ רכב"); continue

        for _, row in df.iterrows():
            vnum = str(row.get(cm["vehicle_num"],"")).strip().replace(".0","")
            if not vnum or vnum in ["nan","","None"]: continue

            start_raw = str(row.get(cm["start"],"") if cm["start"] else "").strip()
            start = ""
            if start_raw not in ["nan","None",""]:
                try: start = pd.to_datetime(start_raw, dayfirst=True).strftime("%Y-%m-%d")
                except: pass

            model = str(row.get(cm["model"],"") if cm["model"] else "").replace("\xa0"," ").strip()
            cat_r = str(row.get(cm["category"],"") if cm["category"] else model)
            cat   = normalize_cat(cat_r) if cat_r not in ["nan","None",""] else normalize_cat(model)

            try:
                rate_r = row.get(cm["rate_ils"]) if cm["rate_ils"] else None
                rate = int(float(str(rate_r).replace(",",""))) if rate_r and str(rate_r) not in ["nan","None",""] else PRICE_ILS.get(cat,8000)
            except: rate = PRICE_ILS.get(cat,8000)

            try:
                days_r = row.get(cm["days"]) if cm["days"] else None
                days = int(float(str(days_r))) if days_r and str(days_r) not in ["nan","None",""] else 30
            except: days = 30

            try:
                km_r = row.get(cm["km"]) if cm["km"] else None
                km = int(float(str(km_r).replace(",",""))) if km_r and str(km_r) not in ["nan","None",""] else None
            except: km = None

            results.append({
                "company": company, "vehicle_num": vnum, "start": start,
                "contract": str(row.get(cm["contract"],"") if cm["contract"] else "").strip().replace(".0",""),
                "model": model if model not in ["nan","None",""] else "",
                "category": cat, "rate_ils": rate, "days": days, "km": km,
                "driver": str(row.get(cm["driver"],"") if cm["driver"] else "").strip(),
                "branch": str(row.get(cm["branch"],"") if cm["branch"] else "").strip(),
                "status": "Active",
            })
    return results, errors

# ── Excel export ───────────────────────────────────────────────────────────
def make_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, bg="FFFFFF", fg="000000", sz=10, bold=False, align="right", wrap=False):
    cell.font      = Font(name="Arial", size=sz, bold=bold, color=fg)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border    = make_border()

def export_excel(data, fx=3.05):
    wb  = Workbook()
    NAV = "1B2A4A"; BLU = "2E5D9F"; WHT = "FFFFFF"; GRY = "F5F5F5"
    LBL = "D6E4F0"; GRN = "1B5E20"; RED = "B71C1C"
    CO  = {"Shlomo":"E8F5E9","Avis":"FCE4EC","Blue Sky":"E3F2FD","Albar":"FFF9E6"}

    # Sheet 1 — Vehicle Database
    ws = wb.active; ws.title = "Vehicle Database"
    ws.merge_cells("A1:P1")
    c = ws["A1"]; c.value = "🚗  Fleet Management — Noamer יעוץ והשקעות"
    cs(c, bg=NAV, fg=WHT, sz=13, bold=True, align="center"); ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:P2")
    c = ws["A2"]; c.value = f"עודכן: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  סה״כ: {len(data)} רכבים  |  שער: $1 = ₪{fx}"
    cs(c, bg=LBL, fg="333333", sz=9, align="center"); ws.row_dimensions[2].height=16

    hdrs = ["ID","חברה","חוזה","תאריך","ימים","מס׳ רכב","דגם","קטגוריה","ק״מ","עלות ₪","הכנסה $","יום חיוב","נהג","סניף","סטטוס"]
    for ci, h in enumerate(hdrs, 1):
        cs(ws.cell(row=3, column=ci, value=h), bg=BLU, fg=WHT, bold=True, align="center", sz=9)
    ws.row_dimensions[3].height = 26; ws.freeze_panes = "A4"; ws.auto_filter.ref = "A3:O3"

    for ri, v in enumerate(data, 4):
        bg = CO.get(v.get("company",""), "FFFFFF")
        d  = v.get("start","")
        try:    bill_day = datetime.strptime(d, "%Y-%m-%d").day
        except: bill_day = ""
        vals = [v.get("id",""), v.get("company",""), v.get("contract",""), d,
                v.get("days",30), v.get("vehicle_num",""), v.get("model",""), v.get("category",""),
                v.get("km"), v.get("rate_ils",0),
                round(PRICE_USD.get(v.get("category","Other"),2475)),
                bill_day, v.get("driver",""), v.get("branch",""), v.get("status","Active")]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            al = "center" if ci in [1,4,5,9,10,11,12] else "left"
            cs(c, bg=bg, align=al, sz=9)
            if ci == 10: c.number_format = "₪#,##0"
            if ci == 11: c.number_format = "$#,##0"
        ws.row_dimensions[ri].height = 18

    for i, w in enumerate([6,14,14,13,7,13,22,18,10,12,12,11,24,20,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2 — Billing Schedule
    ws2 = wb.create_sheet("Billing Schedule")
    ws2.merge_cells("A1:G1")
    c = ws2["A1"]; c.value = "לוח חיובים חודשי · Monthly Billing Schedule"
    cs(c, bg=NAV, fg=WHT, sz=12, bold=True, align="center"); ws2.row_dimensions[1].height = 26
    for ci, h in enumerate(["יום חיוב","חברה","חוזה","מס׳ רכב","קטגוריה","עלות ₪","ימים"], 1):
        cs(ws2.cell(row=2, column=ci, value=h), bg=BLU, fg=WHT, bold=True, align="center", sz=9)
    ws2.row_dimensions[2].height = 22

    rows_b = []
    for v in data:
        s = v.get("start","")
        if not s: continue
        try:   rows_b.append((datetime.strptime(s,"%Y-%m-%d").day, v))
        except: continue
    rows_b.sort(key=lambda x: x[0])

    for ri, (d, v) in enumerate(rows_b, 3):
        bg = CO.get(v.get("company",""), "FFFFFF")
        cs(ws2.cell(row=ri, column=1, value=d),                  bg=bg, align="center", bold=True)
        cs(ws2.cell(row=ri, column=2, value=v.get("company","")), bg=bg, align="left")
        cs(ws2.cell(row=ri, column=3, value=v.get("contract","")),bg=bg, align="left", fg="555555", sz=9)
        cs(ws2.cell(row=ri, column=4, value=v.get("vehicle_num","")), bg=bg, align="left", sz=9)
        cs(ws2.cell(row=ri, column=5, value=v.get("category","")),bg=bg, align="left", fg="555555", sz=9)
        c6 = ws2.cell(row=ri, column=6, value=v.get("rate_ils",0)); cs(c6, bg=bg, bold=True); c6.number_format="₪#,##0"
        cs(ws2.cell(row=ri, column=7, value=v.get("days",30)),    bg=bg, align="center")
        ws2.row_dimensions[ri].height = 18

    for i, w in enumerate([12,14,15,14,20,14,8], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.auto_filter.ref = "A2:G2"; ws2.freeze_panes = "A3"

    # Sheet 3 — Summary
    ws3 = wb.create_sheet("Summary"); ws3.merge_cells("A1:G1")
    c = ws3["A1"]; c.value = "סיכום לפי חברה · Summary by Company"
    cs(c, bg=NAV, fg=WHT, sz=12, bold=True, align="center"); ws3.row_dimensions[1].height = 26
    for ci, h in enumerate(["חברה","רכבים","הכנסה $","הכנסה ₪","עלות ₪","רווח ₪","מרג׳ין"], 1):
        cs(ws3.cell(row=2, column=ci, value=h), bg=BLU, fg=WHT, bold=True, align="center", sz=9)
    for ri, co in enumerate(COMPANIES, 3):
        veh = [v for v in data if v.get("company")==co]
        rev_usd = sum(PRICE_USD.get(v.get("category","Other"),2475) for v in veh)
        rev_ils = round(rev_usd * fx)
        cost    = sum((COSTS_ILS.get(v.get("category",""),{}).get(co) or 0) for v in veh)
        profit  = rev_ils - cost
        margin  = round(profit/rev_ils*100) if rev_ils else 0
        bg = CO.get(co,"FFFFFF")
        for ci, val in enumerate([co, len(veh), rev_usd, rev_ils, cost, profit, f"{margin}%"], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            cs(c, bg=bg, align="right" if ci > 1 else "left", bold=(ci==1))
            if ci in [3]: c.number_format = "$#,##0"
            if ci in [4,5,6]: c.number_format = "₪#,##0"
        ws3.row_dimensions[ri].height = 20
    for i, w in enumerate([16,10,14,14,14,14,10], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO(); wb.save(out); out.seek(0)
    return out

# ── UI CSS ─────────────────────────────────────────────────────────────────
def inject_css():
    st.markdown("""
    <style>
    #MainMenu{visibility:hidden} footer{visibility:hidden}
    .block-container{padding:1.5rem 2rem}
    .metric-card{background:#f8f9fa;border-radius:10px;padding:1rem 1.2rem;
                 border:1px solid #e9ecef;text-align:center;height:100%}
    .metric-val{font-size:1.7rem;font-weight:700;color:#1B2A4A;line-height:1.1}
    .metric-lbl{font-size:0.72rem;color:#6c757d;margin-top:3px;text-transform:uppercase;letter-spacing:.04em}
    .metric-sub{font-size:0.78rem;color:#1D9E75;margin-top:2px;font-weight:500}
    .role-badge-admin{background:#1B2A4A;color:white;padding:2px 10px;border-radius:10px;font-size:11px}
    .role-badge-editor{background:#1565C0;color:white;padding:2px 10px;border-radius:10px;font-size:11px}
    .role-badge-viewer{background:#37474F;color:white;padding:2px 10px;border-radius:10px;font-size:11px}
    .page-title{font-size:1.3rem;font-weight:600;color:#1B2A4A;margin-bottom:1rem;padding-bottom:.5rem;border-bottom:2px solid #E3F2FD}
    div[data-testid="stDataFrame"]{border-radius:8px;overflow:hidden}
    </style>
    """, unsafe_allow_html=True)

def metric(col, val, label, sub=""):
    with col:
        sub_html = f'<div class="metric-sub">{sub}</div>' if sub else ""
        st.markdown(f"""<div class="metric-card">
            <div class="metric-val">{val}</div>
            <div class="metric-lbl">{label}</div>{sub_html}
        </div>""", unsafe_allow_html=True)

# ── Pages ──────────────────────────────────────────────────────────────────
def page_dashboard(db, fx):
    st.markdown('<div class="page-title">📊 דשבורד · Dashboard</div>', unsafe_allow_html=True)
    total      = len(db)
    rev_usd    = sum(PRICE_USD.get(v.get("category","Other"),2475) for v in db)
    rev_ils    = round(rev_usd * fx)
    cost_ils   = sum((COSTS_ILS.get(v.get("category",""),{}).get(v.get("company","")) or 0) for v in db)
    profit     = rev_ils - cost_ils
    no_date    = sum(1 for v in db if not v.get("start",""))

    c1,c2,c3,c4,c5 = st.columns(5)
    metric(c1, total,         "רכבים פעילים",   "")
    metric(c2, f"${rev_usd:,}","הכנסה חודשית",  f"₪{rev_ils:,}")
    metric(c3, f"₪{cost_ils:,}","עלות חודשית",  "")
    metric(c4, f"₪{profit:,}",  "רווח גולמי",   f"{round(profit/rev_ils*100) if rev_ils else 0}%")
    metric(c5, no_date,        "ללא תאריך חיוב", "")

    st.markdown("<br>", unsafe_allow_html=True)
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("##### פירוט לפי חברה")
        rows = []
        for co in COMPANIES:
            veh = [v for v in db if v.get("company")==co]
            if not veh: continue
            ru  = sum(PRICE_USD.get(v.get("category","Other"),2475) for v in veh)
            ri  = round(ru*fx)
            co2 = sum((COSTS_ILS.get(v.get("category",""),{}).get(co) or 0) for v in veh)
            rows.append({"חברה":co,"רכבים":len(veh),"הכנסה $":f"${ru:,}",
                         "הכנסה ₪":f"₪{ri:,}","עלות ₪":f"₪{co2:,}",
                         "רווח ₪":f"₪{ri-co2:,}"})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with col_b:
        st.markdown("##### התפלגות קטגוריות")
        cc = {}
        for v in db: cc[v.get("category","Other")] = cc.get(v.get("category","Other"),0)+1
        df_c = pd.DataFrame({"קטגוריה":list(cc.keys()),"רכבים":list(cc.values())}).sort_values("רכבים",ascending=False)
        st.dataframe(df_c, use_container_width=True, hide_index=True)

    st.markdown("##### ⏰ חיובים ב-10 הימים הקרובים")
    today   = date.today()
    upcoming= []
    for v in db:
        s = v.get("start","")
        if not s: continue
        try:
            d  = datetime.strptime(s,"%Y-%m-%d").day
            bd = date(today.year, today.month, d)
            if bd < today:
                m2 = today.month % 12 + 1
                y2 = today.year + (1 if today.month==12 else 0)
                bd = date(y2, m2, d)
            diff = (bd - today).days
            if 0 <= diff <= 10:
                upcoming.append({"ימים":diff,"תאריך":bd.strftime("%d/%m/%Y"),
                                 "חברה":v.get("company",""),"רכב":v.get("vehicle_num",""),
                                 "קטגוריה":v.get("category",""),"עלות ₪":f"₪{v.get('rate_ils',0):,}"})
        except: continue
    if upcoming:
        st.dataframe(pd.DataFrame(upcoming).sort_values("ימים"), use_container_width=True, hide_index=True)
    else:
        st.info("אין חיובים צפויים ב-10 הימים הקרובים")

def page_upload(db):
    st.markdown('<div class="page-title">📂 העלאת קובץ ספק</div>', unsafe_allow_html=True)
    st.info("המערכת תזהה את הספק אוטומטית ותוסיף רק רכבים חדשים. רכבים קיימים יישארו ללא שינוי.")

    col1, col2 = st.columns([2,1])
    with col1: uploaded = st.file_uploader("בחר קובץ Excel מהספק", type=["xlsx","xls"])
    with col2:
        co_sel = st.selectbox("ספק (אם לא זוהה אוטומטית)", ["זיהוי אוטומטי"]+COMPANIES)
        co_override = None if co_sel=="זיהוי אוטומטי" else co_sel

    if not uploaded: return
    parsed, errors = parse_file(uploaded, co_override)
    for e in errors: st.warning(e)
    if not parsed: st.error("לא נמצאו רכבים בקובץ"); return

    existing = {v.get("vehicle_num","") for v in db}
    new_v    = [v for v in parsed if v.get("vehicle_num","") not in existing]
    old_v    = len(parsed) - len(new_v)

    c1,c2,c3 = st.columns(3)
    c1.metric("רכבים בקובץ", len(parsed))
    c2.metric("רכבים חדשים", len(new_v), delta=f"+{len(new_v)}" if new_v else "0")
    c3.metric("כבר קיימים (מתעלמים)", old_v)

    if new_v:
        st.markdown("##### רכבים שיתווספו:")
        st.dataframe(
            pd.DataFrame(new_v)[["company","vehicle_num","start","category","rate_ils","model"]].rename(
                columns={"company":"חברה","vehicle_num":"מס׳ רכב","start":"תאריך",
                         "category":"קטגוריה","rate_ils":"עלות ₪","model":"דגם"}),
            use_container_width=True, hide_index=True)
        if st.button("✅ אשר והוסף לדאטאבייס", type="primary"):
            nid = next_id(db)
            for i,v in enumerate(new_v): v["id"] = nid+i
            db.extend(new_v); save_db(db)
            st.success(f"✅ נוספו {len(new_v)} רכבים חדשים!")
            st.balloons()
    else:
        st.success("כל הרכבים בקובץ כבר קיימים — אין מה להוסיף.")

def page_add(db):
    st.markdown('<div class="page-title">➕ הוספת רכב ידנית</div>', unsafe_allow_html=True)
    with st.form("add_form"):
        c1,c2,c3 = st.columns(3)
        with c1:
            company = st.selectbox("חברת השכרה *", COMPANIES)
            vnum    = st.text_input("מס׳ רכב *")
            contract= st.text_input("מס׳ חוזה")
        with c2:
            model   = st.text_input("דגם רכב")
            category= st.selectbox("קטגוריה *", CATEGORIES)
            start   = st.date_input("תאריך התחלה")
        with c3:
            rate    = st.number_input("עלות ₪/חודש", min_value=0, value=PRICE_ILS.get(category,8000), step=100)
            days    = st.number_input("ימי חוזה", min_value=1, value=30)
            km      = st.number_input("ק״מ יציאה", min_value=0, value=0, step=100)
        driver  = st.text_input("נהג"); branch = st.text_input("סניף")
        ok = st.form_submit_button("➕ הוסף רכב", type="primary")

    if ok:
        if not vnum: st.error("מס׳ רכב חובה")
        elif any(v.get("vehicle_num","")==vnum for v in db): st.warning(f"רכב {vnum} כבר קיים")
        else:
            db.append({"id":next_id(db),"company":company,"contract":contract,
                       "start":start.strftime("%Y-%m-%d"),"days":days,"vehicle_num":vnum,
                       "model":model,"category":category,"km":km if km>0 else None,
                       "rate_ils":rate,"driver":driver,"branch":branch,"status":"Active"})
            save_db(db); st.success(f"✅ רכב {vnum} נוסף!")

def page_search(db, role):
    st.markdown('<div class="page-title">🔍 חיפוש ועריכה</div>', unsafe_allow_html=True)
    c1,c2,c3 = st.columns(3)
    txt  = c1.text_input("חיפוש (מס׳ רכב / חוזה / דגם)")
    fco  = c2.multiselect("חברה", COMPANIES, default=COMPANIES)
    fcat = c3.multiselect("קטגוריה", CATEGORIES, default=CATEGORIES)

    filtered = [v for v in db
                if (not txt or any(txt.lower() in str(v.get(k,"")).lower() for k in ["vehicle_num","contract","model"]))
                and v.get("company","") in fco
                and v.get("category","") in fcat]

    st.markdown(f"**{len(filtered)} רכבים**")
    if filtered:
        df_s = pd.DataFrame(filtered)[["id","company","vehicle_num","start","category","model","rate_ils","status"]].copy()
        df_s.columns = ["ID","חברה","מס׳ רכב","תאריך","קטגוריה","דגם","עלות ₪","סטטוס"]
        st.dataframe(df_s, use_container_width=True, hide_index=True)

    if role in ["admin","editor"]:
        st.markdown("---"); st.markdown("##### עדכון סטטוס רכב")
        ca,cb,cc = st.columns(3)
        vnum_e = ca.text_input("מס׳ רכב לעדכון")
        status_e = cb.selectbox("סטטוס חדש", ["Active","Returned","Extended","Cancelled"])
        cc.markdown("<br>", unsafe_allow_html=True)
        if cc.button("עדכן"):
            updated = False
            for v in db:
                if v.get("vehicle_num","")==vnum_e: v["status"]=status_e; updated=True
            if updated: save_db(db); st.success(f"✅ עודכן!")
            else: st.error("רכב לא נמצא")

def page_export(db, fx):
    st.markdown('<div class="page-title">📤 יצוא Excel</div>', unsafe_allow_html=True)
    st.info("הקובץ יכיל 3 גיליונות: Vehicle Database, Billing Schedule, Summary")
    c1,c2 = st.columns(2)
    with c1: efx = st.number_input("שער חליפין", value=fx, step=0.01, format="%.2f")
    with c2: fco = st.multiselect("סינון חברות (ריק = הכל)", COMPANIES)
    data_e = db if not fco else [v for v in db if v.get("company","") in fco]

    ca,cb,cc = st.columns(3)
    ca.metric("רכבים בייצוא", len(data_e))
    cb.metric("הכנסה $", f"${sum(PRICE_USD.get(v.get('category','Other'),2475) for v in data_e):,}")
    cc.metric("עלות ₪", f"₪{sum(v.get('rate_ils',0) for v in data_e):,}")

    if st.button("📥 צור קובץ Excel", type="primary"):
        with st.spinner("מייצר..."):
            excel_bytes = export_excel(data_e, efx)
        st.download_button("⬇️ הורד Excel",data=excel_bytes,
            file_name=f"Fleet_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def page_pnl(db, fx):
    st.markdown('<div class="page-title">💰 רווח והפסד · P&L Analysis</div>', unsafe_allow_html=True)

    # ── Settings bar ────────────────────────────────────────────────
    col_fx, col_info = st.columns([1,3])
    with col_fx:
        fx_use = st.number_input("שער חליפין $→₪", value=fx, step=0.01, format="%.2f", key="pnl_fx")
    with col_info:
        st.markdown("<br>", unsafe_allow_html=True)
        st.info("הכנסה = מחירון USD × שער חליפין  |  עלות = מחיר ששולם לחברת ההשכרה (₪)")

    # ── Grand totals ──────────────────────────────────────────────────
    total_rev   = 0
    total_cost  = 0
    total_veh   = len(db)

    for v in db:
        cat = v.get("category", "Other")
        co  = v.get("company", "")
        rev  = PRICE_USD.get(cat, 2475) * fx_use
        cost = COSTS_ILS.get(cat, {}).get(co) or 0
        total_rev  += rev
        total_cost += cost

    total_profit = total_rev - total_cost
    margin       = round(total_profit / total_rev * 100, 1) if total_rev else 0

    c1,c2,c3,c4 = st.columns(4)
    def kpi(col, val, label, color="#1B2A4A"):
        with col:
            st.markdown(f"""<div class="metric-card">
                <div class="metric-val" style="color:{color}">{val}</div>
                <div class="metric-lbl">{label}</div>
            </div>""", unsafe_allow_html=True)

    kpi(c1, f"₪{total_rev:,.0f}",    "הכנסה חודשית כוללת", "#1D9E75")
    kpi(c2, f"₪{total_cost:,.0f}",   "עלות חודשית כוללת",  "#C62828")
    kpi(c3, f"₪{total_profit:,.0f}", "רווח גולמי",          "#1D9E75" if total_profit >= 0 else "#C62828")
    kpi(c4, f"{margin}%",            "מרג׳ין",               "#1D9E75" if margin >= 0 else "#C62828")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── P&L by Company ────────────────────────────────────────────────
    st.markdown("#### לפי חברת השכרה")
    co_rows = []
    for co in COMPANIES:
        veh  = [v for v in db if v.get("company") == co]
        if not veh: continue
        rev  = sum(PRICE_USD.get(v.get("category","Other"), 2475) * fx_use for v in veh)
        cost = sum(COSTS_ILS.get(v.get("category",""), {}).get(co) or 0 for v in veh)
        prof = rev - cost
        mgn  = round(prof / rev * 100, 1) if rev else 0
        co_rows.append({
            "חברה":        co,
            "רכבים":       len(veh),
            "הכנסה ₪":     f"₪{rev:,.0f}",
            "עלות ₪":      f"₪{cost:,.0f}",
            "רווח ₪":      f"₪{prof:,.0f}",
            "מרג׳ין":      f"{mgn}%",
            "הכנסה שנתית": f"₪{rev*12:,.0f}",
            "רווח שנתי":   f"₪{prof*12:,.0f}",
        })
    st.dataframe(pd.DataFrame(co_rows), use_container_width=True, hide_index=True)

    # ── P&L by Category ───────────────────────────────────────────────
    st.markdown("#### לפי קטגוריית רכב")
    cat_rows = []
    for cat in CATEGORIES:
        veh  = [v for v in db if v.get("category") == cat]
        if not veh: continue
        rev  = sum(PRICE_USD.get(cat, 2475) * fx_use for _ in veh)
        cost = sum(COSTS_ILS.get(cat, {}).get(v.get("company","")) or 0 for v in veh)
        prof = rev - cost
        mgn  = round(prof / rev * 100, 1) if rev else 0
        price_usd = PRICE_USD.get(cat, 2475)
        co_breakdown = {}
        for co in COMPANIES:
            n = sum(1 for v in veh if v.get("company") == co)
            if n: co_breakdown[co] = n
        cat_rows.append({
            "קטגוריה":       cat,
            "רכבים":         len(veh),
            "מחיר $ לרכב":   f"${price_usd:,}",
            "הכנסה ₪":       f"₪{rev:,.0f}",
            "עלות ₪":        f"₪{cost:,.0f}",
            "רווח ₪":        f"₪{prof:,.0f}",
            "מרג׳ין":        f"{mgn}%",
            "חברות":         " | ".join(f"{co}({n})" for co, n in co_breakdown.items()),
        })
    st.dataframe(pd.DataFrame(cat_rows), use_container_width=True, hide_index=True)

    # ── Annual projection ─────────────────────────────────────────────
    st.markdown("#### תחזית שנתית")
    c1, c2, c3 = st.columns(3)
    kpi(c1, f"₪{total_rev*12:,.0f}",    "הכנסה שנתית צפויה",  "#1D9E75")
    kpi(c2, f"₪{total_cost*12:,.0f}",   "עלות שנתית צפויה",   "#C62828")
    kpi(c3, f"₪{total_profit*12:,.0f}", "רווח שנתי צפוי",     "#1D9E75" if total_profit >= 0 else "#C62828")

    st.markdown("<br>", unsafe_allow_html=True)
    st.caption(f"* תחזית מבוססת על {total_veh} רכבים פעילים × 12 חודשים  |  שער: $1 = ₪{fx_use}")

    # ── Vehicles without cost data ────────────────────────────────────
    no_cost = [v for v in db if not COSTS_ILS.get(v.get("category",""), {}).get(v.get("company",""))]
    if no_cost:
        with st.expander(f"⚠️ {len(no_cost)} רכבים ללא נתוני עלות — לחץ לפירוט"):
            nc_df = pd.DataFrame(no_cost)[["company","vehicle_num","category","model"]].rename(
                columns={"company":"חברה","vehicle_num":"מס׳ רכב","category":"קטגוריה","model":"דגם"})
            st.dataframe(nc_df, use_container_width=True, hide_index=True)
            st.info("כדי להוסיף עלויות — עדכן את מילון COSTS_ILS בקוד")


def page_users():
    st.markdown('<div class="page-title">👥 ניהול משתמשים</div>', unsafe_allow_html=True)
    users = get_users()
    rows  = [{"שם":u["name"],"אימייל":u["email"],"תפקיד":u["role"]} for u in users.values()]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    st.info("כדי להוסיף/להסיר משתמשים — ערוך את קובץ **secrets.toml** ב-Streamlit Cloud")
    st.code("""[users.new_user]
name         = "שם המשתמש"
email        = "email@noamer.co"
password_hash = "הכנסה הצפנה כאן"
role         = "editor"   # admin / editor / viewer""", language="toml")

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    inject_css()

    if "user" not in st.session_state:
        login_page(); return

    user = st.session_state["user"]
    role = user["role"]
    db   = load_db()
    fx   = st.secrets.get("settings", {}).get("fx_rate", 3.05)

    # Sidebar
    with st.sidebar:
        st.markdown(f"### 🚗 Fleet Management")
        st.markdown(f"**{user['name']}**")
        role_html = f'<span class="role-badge-{role}">{role.upper()}</span>'
        st.markdown(role_html, unsafe_allow_html=True)
        st.markdown(f"<small style='color:#999'>{user['email']}</small>", unsafe_allow_html=True)
        st.divider()

        pages_all    = ["📊 דשבורד","📂 העלאת קובץ ספק","➕ הוספת רכב","🔍 חיפוש ועריכה","💰 רווח והפסד","📤 יצוא Excel","👥 משתמשים"]
        pages_editor = ["📊 דשבורד","📂 העלאת קובץ ספק","➕ הוספת רכב","🔍 חיפוש ועריכה","💰 רווח והפסד","📤 יצוא Excel"]
        pages_viewer = ["📊 דשבורד","🔍 חיפוש ועריכה","💰 רווח והפסד","📤 יצוא Excel"]

        allowed = pages_all if role=="admin" else (pages_editor if role=="editor" else pages_viewer)
        page = st.radio("", allowed, label_visibility="collapsed")

        st.divider()
        st.markdown(f"<small>📊 {len(db)} רכבים במאגר</small>", unsafe_allow_html=True)
        st.markdown(f"<small>💱 שער: $1 = ₪{fx}</small>", unsafe_allow_html=True)
        st.divider()
        if st.button("🚪 התנתק", use_container_width=True): logout()

    if   "דשבורד"          in page: page_dashboard(db, fx)
    elif "העלאת קובץ"      in page: page_upload(db)
    elif "הוספת רכב"       in page: page_add(db)
    elif "חיפוש"           in page: page_search(db, role)
    elif "רווח והפסד"      in page: page_pnl(db, fx)
    elif "יצוא"            in page: page_export(db, fx)
    elif "משתמשים"         in page: page_users()

if __name__ == "__main__":
    main()
